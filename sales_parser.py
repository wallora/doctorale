"""
Parser Excel vendite farmaceutiche (fogli unità / valori).
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

MESI_NOME: dict[str, int] = {
    "gennaio": 1,
    "febbraio": 2,
    "marzo": 3,
    "aprile": 4,
    "maggio": 5,
    "giugno": 6,
    "luglio": 7,
    "agosto": 8,
    "settembre": 9,
    "ottobre": 10,
    "novembre": 11,
    "dicembre": 12,
    "gen": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "mag": 5,
    "giu": 6,
    "lug": 7,
    "ago": 8,
    "set": 9,
    "ott": 10,
    "nov": 11,
    "dic": 12,
}

MESI_LABEL = {
    1: "Gennaio",
    2: "Febbraio",
    3: "Marzo",
    4: "Aprile",
    5: "Maggio",
    6: "Giugno",
    7: "Luglio",
    8: "Agosto",
    9: "Settembre",
    10: "Ottobre",
    11: "Novembre",
    12: "Dicembre",
}

MICROAREA_RE = re.compile(
    r"^[A-ZÀ-Ü][A-ZÀ-Ü0-9\s\.\-]*\s+\d{1,2}$",
    re.IGNORECASE,
)


@dataclass
class SalesRecord:
    metrica: str  # unita | fatturato
    livello: str  # italia | am | informatore | microarea
    informatore: str
    microarea: str
    prodotto: str
    vendita_ap: Optional[float]
    vendita: Optional[float]
    pct_italia: Optional[float]
    crescita_pct: Optional[float]


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("`", "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_product_name(value: str) -> str:
    text = _norm_text(value)
    if text.lower().startswith("totale "):
        rest = text[7:].strip()
        if rest.lower() == "selezione":
            return "Totale selezione"
        return rest
    return text


def _to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def infer_period_from_filename(filename: str) -> tuple[Optional[int], Optional[int]]:
    """Ritorna (anno, mese) se riconoscibili dal nome file."""
    name = unicodedata.normalize("NFKD", filename)
    name = "".join(ch for ch in name if not unicodedata.combining(ch))
    lower = name.lower()

    mese: Optional[int] = None
    for key, num in MESI_NOME.items():
        if re.search(rf"\b{key}\b", lower):
            mese = num
            break

    anno: Optional[int] = None
    m = re.search(r"\b(20\d{2})\b", lower)
    if m:
        anno = int(m.group(1))

    # Pattern 2026-01 / 2026_01 / 2026/01
    m2 = re.search(r"\b(20\d{2})[\-_/](\d{1,2})\b", lower)
    if m2:
        anno = int(m2.group(1))
        mese = int(m2.group(2))

    return anno, mese


def infer_period_from_workbook(file_bytes: bytes) -> tuple[Optional[int], Optional[int]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=5, max_col=8, values_only=True):
                for cell in row:
                    if not cell:
                        continue
                    text = str(cell)
                    m = re.search(r"Periodo\s+(\d{4})\s*/\s*(\d{1,2})", text, re.I)
                    if m:
                        return int(m.group(1)), int(m.group(2))
    finally:
        wb.close()
    return None, None


def suggest_sheets(sheet_names: list[str]) -> tuple[Optional[str], Optional[str]]:
    """Suggerisce (foglio_unita, foglio_fatturato)."""
    unita = None
    fatturato = None
    for name in sheet_names:
        key = name.lower()
        if unita is None and "unit" in key and "isf" in key:
            unita = name
        if fatturato is None and "valor" in key and "isf" in key:
            fatturato = name
    if unita is None:
        for name in sheet_names:
            if "unit" in name.lower():
                unita = name
                break
    if fatturato is None:
        for name in sheet_names:
            if "valor" in name.lower() and "ltm" not in name.lower():
                fatturato = name
                break
    return unita, fatturato


def list_sheet_names(file_bytes: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _find_product_groups(ws: Worksheet, header_row: int = 2) -> list[tuple[str, int]]:
    """Lista (nome_prodotto, colonna_inizio) per gruppi da 4 colonne."""
    groups: list[tuple[str, int]] = []
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        text = _norm_text(value)
        if not text:
            continue
        # ignora etichette periodo
        if text.lower().startswith("periodo"):
            continue
        if text.lower().startswith("totale") or col == 1:
            # prodotti / totale selezione
            if text.lower().startswith("totale"):
                groups.append((_clean_product_name(text), col))
    return groups


def _is_am_person(name: str) -> bool:
    """True per l'Area Manager (non è un informatore)."""
    key = _norm_text(name).lower().replace("`", "")
    key = re.sub(r"\s+", " ", key)
    return key.startswith("ferrari raffaella") or key == "ferrari"


def _is_microarea_label(label: str) -> bool:
    text = _norm_text(label)
    if not text:
        return False
    if text.lower().startswith("totale"):
        return False
    if _is_numeric_label(text):
        return False
    return bool(MICROAREA_RE.match(text)) or bool(
        re.search(r"\b\d{1,2}$", text)
        and re.search(r"[A-Za-zÀ-ü]", text)
        and " " in text
    )


def _is_numeric_label(label: str) -> bool:
    """True se la cella è un numero (es. % Italia letta per sbaglio come etichetta)."""
    text = _norm_text(label).replace("%", "").replace(",", ".")
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def _is_valid_entity_label(label: str) -> bool:
    text = _norm_text(label)
    if not text or _is_numeric_label(text):
        return False
    # serve almeno una lettera
    return bool(re.search(r"[A-Za-zÀ-ü]", text))


def _detect_sheet_layout(ws: Worksheet) -> dict[str, Any]:
    """
    Rileva layout fogli unità/valori.

    Variante A (più recente valori): AM | REP | BRICK | metriche…
    Variante B (unità / valori più vecchi): BRICK/etichetta in col 1 | metriche…
    """
    label_col = 1
    am_col: Optional[int] = None
    rep_col: Optional[int] = None
    header_row = 3
    product_header_row = 2
    data_start_row = 4

    max_col = min(12, (ws.max_column or 1) + 1)
    for r in range(1, 6):
        for c in range(1, max_col):
            h = _norm_text(ws.cell(r, c).value).lower()
            if not h:
                continue
            if h == "am":
                am_col = c
                header_row = r
            elif h == "rep":
                rep_col = c
                header_row = r
            elif h == "brick" or h.startswith("brick"):
                label_col = c
                header_row = r

    if header_row >= 2:
        product_header_row = header_row - 1
        # se la riga sopra non ha "Totale …", prova header_row-1 già ok
        data_start_row = header_row + 1

    # Se non c'è header BRICK, resta col 1 (layout unità classico)
    use_side_cols = am_col is not None or rep_col is not None

    return {
        "label_col": label_col,
        "am_col": am_col,
        "rep_col": rep_col,
        "product_header_row": product_header_row,
        "data_start_row": data_start_row,
        "informatore_from_side_cols": use_side_cols,
    }


def _parse_metric_sheet(
    ws: Worksheet,
    metrica: str,
    label_col: int,
    product_header_row: int,
    data_start_row: int,
    informatore_from_side_cols: bool = False,
    am_col: Optional[int] = None,
    rep_col: Optional[int] = None,
) -> list[SalesRecord]:
    groups = _find_product_groups(ws, header_row=product_header_row)
    if not groups:
        # fallback: a volte i prodotti sono sulla stessa riga del periodo
        groups = _find_product_groups(ws, header_row=max(1, product_header_row - 1))
    if not groups:
        return []

    records: list[SalesRecord] = []
    current_informatore = ""

    max_row = ws.max_row or data_start_row
    for row in range(data_start_row, max_row + 1):
        label = _norm_text(ws.cell(row, label_col).value)
        if not label:
            continue
        # Evita di interpretare metriche numeriche come etichette di riga
        if _is_numeric_label(label):
            continue

        livello = ""
        informatore = ""
        microarea = ""
        lower = label.lower()

        if lower in {"totale italia", "italia"}:
            livello = "italia"
            current_informatore = ""
            informatore = ""
        elif lower in {"totale am", "totale area"}:
            livello = "am"
            # Il nome in colonna AM (es. Ferrari Raffaella) è l'Area Manager,
            # non un informatore: i dati di questa riga sono il totale area.
            current_informatore = ""
            informatore = ""
        elif lower in {"totale rep", "totale informatore"}:
            livello = "informatore"
            if informatore_from_side_cols and rep_col:
                current_informatore = _norm_text(ws.cell(row, rep_col).value)
            else:
                current_informatore = label
            if _is_am_person(current_informatore):
                continue
            informatore = current_informatore
        elif _is_microarea_label(label):
            livello = "microarea"
            microarea = label.upper()
            informatore = current_informatore
        else:
            if not _is_valid_entity_label(label):
                continue
            if _is_am_person(label):
                # Riga AM scambiata per informatore: ignora
                continue
            # riga informatore (es. "aurello", "AURELLO")
            livello = "informatore"
            current_informatore = label
            informatore = current_informatore

        for prodotto, start_col in groups:
            vendita_ap = _to_float(ws.cell(row, start_col).value)
            vendita = _to_float(ws.cell(row, start_col + 1).value)
            pct_italia = _to_float(ws.cell(row, start_col + 2).value)
            crescita = _to_float(ws.cell(row, start_col + 3).value)

            # salta gruppi totalmente vuoti
            if all(v is None for v in (vendita_ap, vendita, pct_italia, crescita)):
                continue

            records.append(
                SalesRecord(
                    metrica=metrica,
                    livello=livello,
                    informatore=informatore,
                    microarea=microarea,
                    prodotto=prodotto,
                    vendita_ap=vendita_ap,
                    vendita=vendita,
                    pct_italia=pct_italia,
                    crescita_pct=crescita,
                )
            )

    return records


def parse_unita_sheet(ws: Worksheet) -> list[SalesRecord]:
    layout = _detect_sheet_layout(ws)
    return _parse_metric_sheet(
        ws,
        metrica="unita",
        label_col=layout["label_col"],
        product_header_row=layout["product_header_row"],
        data_start_row=layout["data_start_row"],
        informatore_from_side_cols=layout["informatore_from_side_cols"],
        am_col=layout["am_col"],
        rep_col=layout["rep_col"],
    )


def parse_fatturato_sheet(ws: Worksheet) -> list[SalesRecord]:
    layout = _detect_sheet_layout(ws)
    return _parse_metric_sheet(
        ws,
        metrica="fatturato",
        label_col=layout["label_col"],
        product_header_row=layout["product_header_row"],
        data_start_row=layout["data_start_row"],
        informatore_from_side_cols=layout["informatore_from_side_cols"],
        am_col=layout["am_col"],
        rep_col=layout["rep_col"],
    )


def _prefer_informatore_name(a: str, b: str) -> str:
    """Preferisce il nominativo più completo (es. cognome+nome vs solo cognome)."""
    a = _norm_text(a)
    b = _norm_text(b)
    if not a:
        return b
    if not b:
        return a
    a_parts = a.split()
    b_parts = b.split()
    if len(b_parts) != len(a_parts):
        return b if len(b_parts) > len(a_parts) else a
    return b if len(b) >= len(a) else a


def unify_informatore_names(records: list[SalesRecord]) -> list[SalesRecord]:
    """
    Allinea i nomi informatore tra fogli unità/fatturato.

    Nei fogli spesso compaiono forme diverse della stessa persona
    (es. "pinasco" vs "Pinasco Jorge Luis"). L'abbinamento principale
    avviene sulle microaree gestite; in fallback sul cognome.
    """
    from collections import defaultdict

    areas_by_inf: dict[str, set[str]] = defaultdict(set)
    for r in records:
        if r.livello != "microarea":
            continue
        inf = _norm_text(r.informatore)
        area = _norm_text(r.microarea).upper()
        if inf and area:
            areas_by_inf[inf].add(area)

    names = list(areas_by_inf.keys())
    parent = {n: n for n in names}

    def find(x: str) -> str:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        preferred = _prefer_informatore_name(ra, rb)
        other = rb if preferred == ra else ra
        parent[other] = preferred
        parent[preferred] = preferred

    # 1) Match per overlap microaree
    for i, left in enumerate(names):
        for right in names[i + 1 :]:
            a_left = areas_by_inf[left]
            a_right = areas_by_inf[right]
            overlap = len(a_left & a_right)
            if overlap == 0:
                continue
            ratio = overlap / min(len(a_left), len(a_right))
            if ratio >= 0.6 or overlap >= 2:
                union(left, right)

    # 2) Fallback cognome corto -> nominativo lungo
    long_names = [n for n in names if " " in n]
    short_names = [n for n in names if " " not in n]
    for short in short_names:
        short_key = short.lower()
        matches = [
            long
            for long in long_names
            if long.lower().startswith(short_key + " ")
            or long.lower().split()[0] == short_key
        ]
        if len(matches) == 1:
            union(short, matches[0])

    canonical_of = {n: find(n) for n in names}

    for r in records:
        inf = _norm_text(r.informatore)
        if inf in canonical_of:
            r.informatore = canonical_of[inf]

    return records


def parse_sales_workbook(
    file_bytes: bytes,
    sheet_unita: str,
    sheet_fatturato: str,
) -> list[SalesRecord]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    try:
        if sheet_unita not in wb.sheetnames:
            raise ValueError(f"Foglio unità non trovato: {sheet_unita}")
        if sheet_fatturato not in wb.sheetnames:
            raise ValueError(f"Foglio fatturato non trovato: {sheet_fatturato}")
        records = []
        records.extend(parse_unita_sheet(wb[sheet_unita]))
        records.extend(parse_fatturato_sheet(wb[sheet_fatturato]))
        return unify_informatore_names(records)
    finally:
        wb.close()
